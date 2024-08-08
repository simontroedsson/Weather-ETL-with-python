import os
from datetime import time

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from termcolor import colored

file_name = "väderdata.xlsx"


def style_excel_sheet(sheet):
    # Sätt headings till bold och ändra font size
    for col in range(1, 9):
        sheet[get_column_letter(col) + "1"].font = Font(bold=True, size=14)

    # öka column bredd
    length_of_column_A = len(str(sheet["A2"].value))
    sheet.column_dimensions["A"].width = length_of_column_A
    sheet.column_dimensions["B"].width = length_of_column_A
    sheet.column_dimensions["C"].width = length_of_column_A
    sheet.column_dimensions["D"].width = length_of_column_A
    sheet.column_dimensions["E"].width = length_of_column_A
    sheet.column_dimensions["F"].width = length_of_column_A
    sheet.column_dimensions["G"].width = length_of_column_A
    sheet.column_dimensions["H"].width = length_of_column_A


def data_exist_in_sheet(weather_data_next_24h: list[dict]) -> bool:
    loaded_workbook = load_workbook(file_name)
    loaded_sheet = loaded_workbook["väderData"]
    antal_rader = loaded_sheet.max_row
    excel_data_list = []
    # Hämta alla rader från excel som tupler i en Lista
    for row in loaded_sheet.iter_rows(
        min_row=2, max_col=8, max_row=antal_rader, values_only=True
    ):
        excel_data_list.append((row[4], row[5], row[6]))
    # valde att bara ta med timme,temp,rainOrsnow från varje rad

    # Dela upp raderna i listor, 24st i varje lista
    antal_prognoser = int(len(excel_data_list) / 24)
    start = 0
    list_prognoser = []
    for i in range(1, antal_prognoser + 1):
        end = i * 24
        list_prognoser.append(excel_data_list[start:end])
        start = end

    # Hämta nya datan som man vill jämföra med excel datan
    new_data_list = []
    for v_dict in weather_data_next_24h:
        new_data_list.append(
            (
                v_dict["hour"],
                v_dict["temperature"],
                v_dict["RainOrSnow"],
            )
        )
    # jämför nya data med excel datan
    for excel_prognos in list_prognoser:
        if new_data_list == excel_prognos:
            return True
    return False


def append_weather_data_to_sheet(loaded_sheet, weather_data_next_24h: list[dict]):
    for v_dict in weather_data_next_24h:
        loaded_sheet.append(
            [
                v_dict["created"],
                v_dict["longitude"],
                v_dict["latitude"],
                v_dict["date"],
                v_dict["hour"],
                v_dict["temperature"],
                v_dict["RainOrSnow"],
                v_dict["provider"],
            ]
        )


def send_data_to_excel(weather_data_next_24h: list[dict]):
    provider = weather_data_next_24h[0]["provider"]
    error_text = colored(
        f"Senaste prognos från {provider} finns redan i {file_name}, vänta en stund och pröva igen",
        "light_red",
    )
    success_text = colored(
        f"{file_name} är uppdaterad med den senaste prognosen från {provider}.",
        "light_green",
    )

    if os.path.exists(f"{file_name}"):
        if data_exist_in_sheet(weather_data_next_24h):
            print(error_text)
            return
        else:
            loaded_workbook = load_workbook(file_name)
            loaded_sheet = loaded_workbook["väderData"]
            append_weather_data_to_sheet(loaded_sheet, weather_data_next_24h)
            loaded_workbook.save(file_name)
            print(success_text)
    else:
        new_wb = Workbook()
        sheet = new_wb.active
        sheet.title = "väderData"
        headings = list(weather_data_next_24h[0].keys())
        sheet.append(headings)
        append_weather_data_to_sheet(sheet, weather_data_next_24h)
        style_excel_sheet(sheet)
        new_wb.save(file_name)
        print(success_text)
        new_wb.close()


def print_forecast_from_excel(provider_choice: str):

    if os.path.exists(f"{file_name}"):
        loaded_workbook = load_workbook(file_name)
        loaded_sheet = loaded_workbook["väderData"]
        row_ending = loaded_sheet.max_row
        row_start = row_ending - 23

        number_of_forecasts = int((row_ending - 1) / 24)
        i = 1
        while i <= number_of_forecasts:
            if loaded_sheet[f"H{row_start}"].value == provider_choice:
                break
            else:
                row_ending = row_start - 1
                row_start = row_start - 24
                i += 1

        if i > number_of_forecasts:
            error_text = colored(
                f"Prognos från {provider_choice} fanns inte i dokumentet", "light_red"
            )
            print(error_text)
            return

        datum = loaded_sheet[f"D{row_start}"].value
        print(f"Prognos från {provider_choice} {datum.date()}")
        for row in loaded_sheet.iter_rows(
            min_row=row_start, max_col=8, max_row=row_ending, values_only=True
        ):
            rainOrSnow = "ingen nederbörd"
            if row[6] == True:
                rainOrSnow = "nederbörd"
            timme = time(row[4])
            Ftime = timme.strftime("%H:%M")
            temp = row[5]
            print(f"{Ftime} {temp}".ljust(11), f"grader {rainOrSnow}")
            # format(temp,'.1f')
    else:
        error_text = colored(
            f"{file_name} har inga prognoser, Du måste hämta data först", "light_red"
        )
        print(error_text)
